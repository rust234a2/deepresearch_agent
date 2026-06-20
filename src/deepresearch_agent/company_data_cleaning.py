from __future__ import annotations

import csv
import re
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


MISSING_VALUES = {"", "-", "--", "暂无", "null", "none"}

CORE_COLUMNS = [
    "source_name",
    "legal_name",
    "registration_status",
    "unified_social_credit_code",
    "legal_representative",
    "company_type",
    "registered_capital_amount",
    "registered_capital_currency",
    "registered_capital_original",
    "paid_in_capital_amount",
    "paid_in_capital_currency",
    "paid_in_capital_original",
    "established_date",
    "business_term_start",
    "business_term_end",
    "business_term_indefinite",
    "registered_address",
    "province",
    "city",
    "district",
    "registration_authority",
    "gb_industry_section",
    "gb_industry_division",
    "gb_industry_group",
    "gb_industry_class",
    "enterprise_size",
    "business_scope",
    "aliases",
    "english_name",
    "website",
    "employee_count",
    "employee_count_report_year",
    "latest_annual_report_year",
    "taxpayer_qualification",
]

CONTACT_COLUMNS = [
    "unified_social_credit_code",
    "legal_name",
    "phones",
    "emails",
    "mailing_address",
]

REJECTED_COLUMNS = ["source_name", "matched_name", "reason"]


def normalize_missing(value: object) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).replace("\r", "\n").split())
    if text.casefold() in MISSING_VALUES:
        return ""
    return text


def parse_capital(value: object) -> tuple[str, str, str]:
    original = normalize_missing(value)
    if not original:
        return "", "", ""

    match = re.search(r"([0-9][0-9,]*(?:\.[0-9]+)?)", original)
    if not match:
        return "", "", original
    try:
        amount = Decimal(match.group(1).replace(",", ""))
    except InvalidOperation:
        return "", "", original
    if "万" in original:
        amount *= Decimal(10000)

    if "港元" in original or "港币" in original:
        currency = "HKD"
    elif "美元" in original:
        currency = "USD"
    elif "欧元" in original:
        currency = "EUR"
    else:
        currency = "CNY"
    if amount == amount.to_integral_value():
        normalized_amount = str(int(amount))
    else:
        normalized_amount = format(amount, "f").rstrip("0").rstrip(".")
    return normalized_amount, currency, original


def parse_business_term(value: object) -> tuple[str, str, bool]:
    text = normalize_missing(value)
    if not text:
        return "", "", False
    parts = re.split(r"\s*至\s*", text, maxsplit=1)
    start = normalize_date(parts[0])
    if len(parts) == 1:
        return start, "", False
    end_text = parts[1]
    if "无固定期限" in end_text or "长期" == end_text.strip():
        return start, "", True
    return start, normalize_date(end_text), False


def split_values(value: object) -> list[str]:
    text = normalize_missing(value)
    if not text:
        return []
    values: list[str] = []
    seen: set[str] = set()
    for item in re.split(r"[;；,，\n]+", text):
        normalized = normalize_missing(item)
        key = normalized.casefold()
        if normalized and key not in seen:
            seen.add(key)
            values.append(normalized)
    return values


def normalize_date(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = normalize_missing(value)
    match = re.search(r"\d{4}-\d{1,2}-\d{1,2}", text)
    if not match:
        return text
    year, month, day = match.group(0).split("-")
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def normalize_year(value: object) -> str:
    match = re.search(r"\d{4}", normalize_missing(value))
    return match.group(0) if match else ""


def clean_rows(rows: Iterable[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    companies: list[dict] = []
    contacts: list[dict] = []
    rejected: list[dict] = []
    for row in rows:
        source_name = normalize_missing(row.get("原文件导入名称"))
        legal_name = normalize_missing(row.get("系统匹配企业名称"))
        if legal_name == "未匹配到相关企业":
            rejected.append(
                {"source_name": source_name, "matched_name": legal_name, "reason": "unmatched"}
            )
            continue

        registration_id = normalize_missing(row.get("统一社会信用代码"))
        if not legal_name or not registration_id:
            rejected.append(
                {
                    "source_name": source_name,
                    "matched_name": legal_name,
                    "reason": "missing_identity",
                }
            )
            continue

        registered_amount, registered_currency, registered_original = parse_capital(
            row.get("注册资本")
        )
        paid_amount, paid_currency, paid_original = parse_capital(row.get("实缴资本"))
        term_start, term_end, term_indefinite = parse_business_term(row.get("营业期限"))
        aliases = split_values(row.get("曾用名"))
        phones = split_values(
            ";".join(
                filter(
                    None,
                    [normalize_missing(row.get("电话")), normalize_missing(row.get("更多电话"))],
                )
            )
        )
        emails = split_values(
            ";".join(
                filter(
                    None,
                    [normalize_missing(row.get("邮箱")), normalize_missing(row.get("更多邮箱"))],
                )
            )
        )

        companies.append(
            {
                "source_name": source_name,
                "legal_name": legal_name,
                "registration_status": normalize_missing(row.get("登记状态")),
                "unified_social_credit_code": registration_id,
                "legal_representative": normalize_missing(row.get("法定代表人")),
                "company_type": normalize_missing(row.get("企业（机构）类型")),
                "registered_capital_amount": registered_amount,
                "registered_capital_currency": registered_currency,
                "registered_capital_original": registered_original,
                "paid_in_capital_amount": paid_amount,
                "paid_in_capital_currency": paid_currency,
                "paid_in_capital_original": paid_original,
                "established_date": normalize_date(row.get("成立日期")),
                "business_term_start": term_start,
                "business_term_end": term_end,
                "business_term_indefinite": str(term_indefinite).lower(),
                "registered_address": normalize_missing(row.get("企业地址")),
                "province": normalize_missing(row.get("所属省份")),
                "city": normalize_missing(row.get("所属城市")),
                "district": normalize_missing(row.get("所属区县")),
                "registration_authority": normalize_missing(row.get("登记机关")),
                "gb_industry_section": normalize_missing(row.get("国标行业门类")),
                "gb_industry_division": normalize_missing(row.get("国标行业大类")),
                "gb_industry_group": normalize_missing(row.get("国标行业中类")),
                "gb_industry_class": normalize_missing(row.get("国标行业小类")),
                "enterprise_size": normalize_missing(row.get("企业规模")),
                "business_scope": normalize_missing(row.get("经营范围")),
                "aliases": "|".join(aliases),
                "english_name": normalize_missing(row.get("英文名")),
                "website": normalize_missing(row.get("官网")),
                "employee_count": _digits(row.get("参保人数")),
                "employee_count_report_year": normalize_year(row.get("参保人数所属年报")),
                "latest_annual_report_year": normalize_year(row.get("最新年报年份")),
                "taxpayer_qualification": normalize_missing(row.get("纳税人资质")),
            }
        )
        contacts.append(
            {
                "unified_social_credit_code": registration_id,
                "legal_name": legal_name,
                "phones": "|".join(phones),
                "emails": "|".join(emails),
                "mailing_address": normalize_missing(row.get("通信地址")),
            }
        )
    return companies, contacts, rejected


def read_workbook_rows(path: str | Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = [normalize_missing(cell.value) for cell in worksheet[2]]
        rows: list[dict] = []
        for values in worksheet.iter_rows(min_row=3, values_only=True):
            row = {header: value for header, value in zip(headers, values) if header}
            if any(normalize_missing(value) for value in row.values()):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def write_csv(rows: Iterable[dict], columns: list[str], path: str | Path) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _digits(value: object) -> str:
    match = re.search(r"\d+", normalize_missing(value).replace(",", ""))
    return match.group(0) if match else ""
