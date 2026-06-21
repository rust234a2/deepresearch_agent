import csv
import re
import subprocess
import sys
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

from deepresearch_agent.company_data_cleaning import (
    clean_rows,
    normalize_missing,
    parse_business_term,
    parse_capital,
    run_cleaning,
    split_values,
)


def test_normalize_missing_treats_masked_value_as_missing_but_preserves_embedded_separator():
    assert normalize_missing("***") == ""
    assert normalize_missing("许可项目***一般项目") == "许可项目***一般项目"


def test_parse_capital_converts_ten_thousand_yuan():
    assert parse_capital("13,400万元") == ("134000000", "CNY", "13,400万元")


def test_parse_capital_supports_foreign_currency():
    assert parse_capital("500万美元") == ("5000000", "USD", "500万美元")


def test_parse_business_term_handles_indefinite_end():
    assert parse_business_term("1997-01-28 至 无固定期限") == (
        "1997-01-28",
        "",
        True,
    )


def test_split_values_removes_placeholders_and_duplicates():
    assert split_values("a@example.com;a@example.com;-;b@example.com") == [
        "a@example.com",
        "b@example.com",
    ]


def test_clean_rows_separates_matched_and_rejected_records():
    rows = [
        {
            "原文件导入名称": "示例科技",
            "系统匹配企业名称": "示例科技股份有限公司",
            "登记状态": "存续",
            "统一社会信用代码": "91330000123456789X",
            "法定代表人": "张三",
            "企业（机构）类型": "股份有限公司",
            "注册资本": "100万元",
            "实缴资本": "50万元",
            "成立日期": "2020-01-02",
            "营业期限": "2020-01-02 至 无固定期限",
            "企业地址": "浙江省杭州市示例路1号",
            "所属省份": "浙江省",
            "所属城市": "杭州市",
            "所属区县": "西湖区",
            "电话": "0571-12345678",
            "更多电话": "0571-12345678;400-123-4567",
            "邮箱": "info@example.com",
            "更多邮箱": "info@example.com;sales@example.com",
            "国标行业门类": "制造业",
            "国标行业大类": "专用设备制造业",
            "曾用名": "示例设备有限公司,\n示例机械有限公司",
            "参保人数": "120",
            "参保人数所属年报": "2025年报",
            "最新年报年份": "2025",
            "通信地址": "-",
        },
        {
            "原文件导入名称": "未匹配公司",
            "系统匹配企业名称": "未匹配到相关企业",
        },
    ]

    companies, contacts, rejected = clean_rows(rows)

    assert len(companies) == 1
    assert companies[0]["legal_name"] == "示例科技股份有限公司"
    assert companies[0]["registered_capital_amount"] == "1000000"
    assert companies[0]["business_term_indefinite"] == "true"
    assert companies[0]["aliases"] == "示例设备有限公司|示例机械有限公司"
    assert companies[0]["employee_count_report_year"] == "2025"
    assert contacts == [
        {
            "unified_social_credit_code": "91330000123456789X",
            "legal_name": "示例科技股份有限公司",
            "phones": "0571-12345678|400-123-4567",
            "emails": "info@example.com|sales@example.com",
            "mailing_address": "",
        }
    ]
    assert rejected == [
        {
            "source_name": "未匹配公司",
            "matched_name": "未匹配到相关企业",
            "reason": "unmatched",
        }
    ]


def test_run_cleaning_writes_three_csv_files(tmp_path):
    workbook_path = tmp_path / "source.xlsx"
    output_dir = tmp_path / "cleaned"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["数据使用声明"])
    worksheet.append(
        [
            "原文件导入名称",
            "系统匹配企业名称",
            "登记状态",
            "统一社会信用代码",
            "注册资本",
            "电话",
        ]
    )
    worksheet.append(
        ["示例科技", "示例科技股份有限公司", "存续", "91330000123456789X", "100万元", "123"]
    )
    worksheet.append(["未匹配公司", "未匹配到相关企业", "-", "-", "-", "-"])
    workbook.save(workbook_path)

    summary = run_cleaning(workbook_path, output_dir)

    assert summary == {"input_rows": 2, "companies": 1, "contacts": 1, "rejected": 1}
    for file_name in ("companies.csv", "contacts.csv", "rejected.csv"):
        assert (output_dir / file_name).exists()
    with (output_dir / "companies.csv").open(encoding="utf-8-sig", newline="") as handle:
        companies = list(csv.DictReader(handle))
    assert companies[0]["legal_name"] == "示例科技股份有限公司"


def test_run_cleaning_handles_incorrect_workbook_dimension(tmp_path):
    workbook_path = tmp_path / "source.xlsx"
    rewritten_path = tmp_path / "rewritten.xlsx"
    output_dir = tmp_path / "cleaned"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["数据使用声明"])
    worksheet.append(["原文件导入名称", "系统匹配企业名称", "统一社会信用代码"])
    worksheet.append(["示例科技", "示例科技股份有限公司", "91330000123456789X"])
    workbook.save(workbook_path)

    with ZipFile(workbook_path) as source, ZipFile(rewritten_path, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                content = re.sub(br'<dimension ref="[^"]+"\s*/>', b'<dimension ref="A1"/>', content)
            target.writestr(item, content)

    summary = run_cleaning(rewritten_path, output_dir)

    assert summary["input_rows"] == 1
    assert summary["companies"] == 1


def test_cleaning_script_uses_project_source_when_run_from_other_directory(tmp_path):
    project_root = Path(__file__).parents[1]
    workbook_path = tmp_path / "source.xlsx"
    output_dir = tmp_path / "output"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["数据使用声明"])
    worksheet.append(
        [
            "原文件导入名称",
            "系统匹配企业名称",
            "统一社会信用代码",
            "营业期限",
        ]
    )
    worksheet.append(
        ["示例科技", "示例科技股份有限公司", "91330000123456789X", "*** 至 无固定期限"]
    )
    workbook.save(workbook_path)

    result = subprocess.run(
        [
            sys.executable,
            str(project_root / "scripts" / "clean_qcc_company_data.py"),
            "--input",
            str(workbook_path),
            "--output-dir",
            str(output_dir),
        ],
        cwd=tmp_path,
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    with (output_dir / "companies.csv").open(encoding="utf-8-sig", newline="") as handle:
        company = next(csv.DictReader(handle))
    assert company["business_term_start"] == ""
